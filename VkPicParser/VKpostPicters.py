import vk
import os
from ast import literal_eval
import requests
import json
import time
from openpyxl import Workbook, load_workbook
import datetime
import random
import shutil



# def get_open_fds():
#     ''' return the number of open file descriptors for current process .. warning: will only work on UNIX-like os-es. '''
#     import subprocess
#     import os
#     pid = os.getpid()
#     procs = subprocess.check_output( [ "lsof", '-w', '-Ff', "-p", str( pid ) ] )
#     nprocs = len( filter( lambda s: s and s[ 0 ] == 'f' and s[1: ].isdigit(), procs.split( '\n' ) ) )
#     return nprocs


def timer():
    x = datetime.datetime.now()
    td = int(str(x).split("-")[2].split(" ")[1].split(":")[0])
    td = 24 - td
    return td * 60 * 60


def mdir():
    mypath = "Uploaded_Pictures"
    if not os.path.isdir(mypath):
        os.makedirs(mypath)


def take_name_from_excel(file, gif_name):
    wb = load_workbook(file)

    ws = wb.active
    first_column = ws['G']
    for index, i in enumerate(first_column):
        if i.value == gif_name:
            break
    name = ws['E'][index].value
    # name = str(name)
    return name

    # print(first_column)


def list_files_folder(path):
    lst = os.listdir(path)
    lst_new = []
    if len(lst) > 0:
        for i in lst:
            if i[-3:] == "jpg":
                lst_new.append(i)
        return lst_new
    else:
        return lst_new


def remove(path):
    """ param <path> could either be relative or absolute. """
    if os.path.isfile(path):
        os.remove(path)  # remove the file
    elif os.path.isdir(path):
        shutil.rmtree(path)  # remove dir and all contains
    else:
        raise ValueError("file {} is not a file or dir.".format(path))


def main():
    try:
        data = ""
        path = 'params.txt'
        with open(path, 'r') as f:
            for lines in f:
                data = data + lines
        d = (literal_eval(data))

        path_folder = "Output_Pictures\\"
        destination_folder = "Uploaded_Pictures\\"
        min_shift = d["min_shift"]
        token = d["token_post"]
        owner_id = d["owner_id"]
        token_parser = d["token_parser"]

    except Exception as e:
        print("read from params.txt", e)
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) +"Error on reads param.txt "+ "   " + str(e) + "\n")
        file.close()








    try:
        session = vk.AuthSession(access_token=token)
        vk_api = vk.API(session)
        posts = vk_api.wall.get(owner_id=owner_id * (-1),
                                count=1000, filter="postponed", v=5.101)

        posts = posts["items"][-1]

        # time=(posts["items"][0]["attachments"][0]["doc"]['date'])
        time_ = (posts["date"])
        print("Last postponed post ...", str(datetime.datetime.fromtimestamp(time_)))
        start_shifttime = int(time_)



    except Exception as e:
        x = datetime.datetime.now()
        start_shifttime = datetime.datetime.timestamp(x)
        print("Postponed time is not find, get present time ...", str(datetime.datetime.fromtimestamp(start_shifttime)))
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + "Time difinition error " + "   " + str(e) + "\n")
        file.close()












    if len(list_files_folder(path_folder)) > 0:

        error = 0
        y = start_shifttime + (min_shift * 60)
        n_posts = 1



        while len(list_files_folder(path_folder)) > 0:
            try:
                step = random.randint(7, 15)



                file_name = list_files_folder(path_folder)[0]

                message = take_name_from_excel("VK_list.xlsx", file_name)

                pic_path = path_folder + file_name
                os.rename(pic_path, destination_folder + file_name)
                # os.close(destination_folder + file_name)

                pic_path = destination_folder + file_name
                # print(pic_path)

                session = vk.AuthSession(access_token=token)
                vk_api = vk.API(session)
                result = vk_api.docs.getWallUploadServer(v=5.102)['upload_url']

                img = {'file': (pic_path, open(r'' + pic_path, 'rb'))}

                upload_url = result
                # print(upload_url)
                response = requests.post(upload_url, files=img)

                result = json.loads(response.text)
                result = vk_api.docs.save(file=result['file'], v=5.102)

                attachments = ("doc" + str(result['doc']["owner_id"]) + "_" + str(result['doc']["id"]))

                print(str(datetime.datetime.fromtimestamp(y)))
                print(message)
                print(file_name)

                upload = vk_api.wall.post(message=message, publish_date=y, owner_id=-1 * (owner_id), from_group=1,
                                          attachments=attachments, v=5.102)

                print("Picter  was posted {}, post number {}".format(upload,n_posts))
                y = y + (min_shift * 60)
                time.sleep(step * 60)
                error = 0
                n_posts = n_posts + 1
                if n_posts > 25:
                    time.sleep(timer())
                    n_posts=1


            except Exception as er:
                try:
                    shutil.copy(pic_path, destination_folder + file_name)
                except:
                    pass

                error = error + 1

                print(er)
                with open('log.txt', "a") as file:
                    file.write(str(datetime.datetime.now()) + "   " + str(er) + "\n")
                file.close()
                if error > 3:
                    break


    else:
        print("Pictures folder is empty")

if __name__ == "__main__":
    start = datetime.datetime.now()

    main()

    end = datetime.datetime.now()
    print("Duration: {}".format(end - start))
    t
