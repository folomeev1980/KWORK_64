from requests import get
import vk
import os
from ast import literal_eval
import requests
import json
import time
from openpyxl import Workbook, load_workbook
import datetime
import random


def get_captcha_key():
    with open('captcha_key.txt','rt') as f:
        return f.read().replace('\n','')

RUCAPTCHA_KEY = get_captcha_key()
print()

def captcha_solve(captcha_link):
    from python_rucaptcha import ImageCaptcha,RuCaptchaControl
    answer = RuCaptchaControl.RuCaptchaControl(
        rucaptcha_key=RUCAPTCHA_KEY
    ).additional_methods(action="getbalance")
    print("Checking Rucaptcha balance: ", answer['serverAnswer'])
    image_link = captcha_link
    user_answer = ImageCaptcha.ImageCaptcha(rucaptcha_key=RUCAPTCHA_KEY).captcha_handler(captcha_link=image_link)

    if not user_answer['error']:
        # решение капчи
        print(user_answer['captchaSolve'])
        print(user_answer['taskId'])
        return user_answer['captchaSolve']
    elif user_answer['error']:
        # Тело ошибки, если есть
        print(user_answer['errorBody']['text'])
        print(user_answer['errorBody']['id'])
        return None



def captcha_handler(captcha):
    """ При возникновении капчи вызывается эта функция и ей передается объект
        капчи. Через метод get_url можно получить ссылку на изображение.
        Через метод try_again можно попытаться отправить запрос с кодом капчи
    """
    res = captcha_solve(captcha.get_url())

    if res == None:
        print("Captcha solve error. Please, write manually")
        key = input("Enter captcha code {0}: ".format(captcha.get_url())).strip()
    else:
        print("Captcha solved. Key is: ",res)
        key = res

    # Пробуем снова отправить запрос с капчей
    return captcha.try_again(key)



def photo_video_choice(post):
    post_dic = {}

    if post['attachments'][0]["type"] == "photo":
        # print(post)
        photo = dict(post['attachments'][0]["photo"]["sizes"][-1])
        # post_dic['video'] = photo["url"]
        post_dic['video'] = None
        return post_dic
    else:
        # print("ok")
        # print()
        # video = dict(post['attachments'][0]["doc"]["url"])
        # print(post['attachments'][0]["doc"]["url"])
        post_dic['video'] = post['attachments'][0]["doc"]["url"]
        return post_dic

def get_comments(vk_api, owner_id, post_id):
    comments = vk_api.wall.getComments(
        owner_id=owner_id * (-1), post_id=post_id, count=100, need_likes=1, v=5.101)
    comments_number = comments["count"] // 100

    comments = (comments["items"])
    likes = comments[0]["likes"]["count"]
    text = comments[0]["text"]

    for i in range(1, comments_number + 1):
        comments = comments + vk_api.wall.getComments(
            owner_id=owner_id * (-1), post_id=post_id, need_likes=1, count=100, v=5.101, offset=i * 100)["items"]

    for index, comment in enumerate(comments):
        if comment["likes"]["count"] > likes:
            likes = comment["likes"]["count"]
            text = comment["text"]

    return {"text_comment": text, "likes_comment": likes}


def get_group_posts(*args):
    args = args[0]
    print(args)
    vk_api = args[0]
    owner_id = args[1]
    date1 = args[2]

    date2 = args[3]
    namber_comments_of = args[4]
    likes = args[5]
    ads = args[6]

    post_dic = {}
    posts_list = []
    exception_numbers = 0
    date_start = datetime.datetime(date1[0], date1[1], date1[2])
    date_end = datetime.datetime(date2[0], date2[1], date2[2], 23, 59)

    posts = vk_api.wall.get(owner_id=owner_id * (-1),
                            count=100, filter=all, v=5.101)

    count_numbers = posts["count"] // 100

    posts = posts["items"]

    # Получаем первую дату поста

    for i in range(1, count_numbers + 1):
        print("Reading {} block {}".format(i, owner_id))

        temp = vk_api.wall.get(owner_id=owner_id * (-1), count=100, filter=all,
                               v=5.101, offset=i * 100)["items"]

        start_block = datetime.datetime.fromtimestamp(temp[0]["date"])
        end_block = datetime.datetime.fromtimestamp(temp[len(temp) - 1]["date"])

        if start_block > date_start:
            if end_block < date_end:
                posts = posts + temp
                print("posts block --{}-- is writting".format(i * 100))
        else:
            break

        # time.sleep(2)

    for index, post in enumerate(posts):
        try:
            d = datetime.datetime.fromtimestamp(post['date'])
            # print(d,date_start,date_end)
            if d >= date_start:
                if d <= date_end:

                    post_dic["date"] = d
                    post_dic["id"] = post["id"]
                    post_dic["text"] = post["text"]
                    post_dic['marked_as_ads'] = post['marked_as_ads']
                    # video = dict(post['attachments'][0]["doc"]["preview"]["video"])
                    post_dic['video'] = (photo_video_choice(post))['video']
                    # post_dic['filesize'] = (dict(video)['file_size'])
                    post_dic['comments'] = post['comments']["count"]

                    if post_dic['comments'] > 0:
                        dic_ret = get_comments(vk_api, owner_id, post_dic["id"])
                        post_dic['text_comment'] = dic_ret["text_comment"]
                        post_dic['likes_comment'] = dic_ret["likes_comment"]
                    else:
                        post_dic['text_comment'] = ""
                        post_dic['likes_comment'] = ""

                    post_dic['post_likes'] = post['likes']["count"]
                    post_dic['reposts'] = post['reposts']["count"]
                    post_dic['views'] = post['views']["count"]

                    if post_dic["date"] != "":
                        if post_dic['comments'] > namber_comments_of:
                            if post_dic['likes_comment'] > likes:
                                if post_dic['marked_as_ads'] == ads:
                                    if post_dic["video"] != None:
                                        posts_list.append(post_dic)



        except:
            exception_numbers += 1
            # post_dic = {"date": "", "id": "", "text": "", 'marked_as_ads': "", 'video': "",
            #             'comments': "", 'text_comment': "", 'likes_comment': "", 'post_likes': "", 'reposts': "",
            #             'views': ""}

        if index % 10 == 0:
            print("{} posts  {} was parsed".format(owner_id, index))
        post_dic = {}
    print("Exception numbers is {} from {} pages x 100 posts ".format(
        exception_numbers, count_numbers))

    return posts_list


def excel_creator(dic_list):
    print("Start to create xlx file")
    # dir = os.path.abspath(os.curdir)
    list_of_items = []

    book = Workbook()
    sheet = book.active
    # post_dic = {"date": "", "id": "", "text": "", 'marked_as_ads': "", 'video': "",
    # 'filesize': "", 'comments': "", 'text_comment': "", 'post_likes': "",
    # 'likes_comment': "", 'reposts': "", 'views': ""}
    sheet.append(["Дата и время", "Название поста", "Реклама", "Колличество комментов",
                  "Текст лучшего коммента", "Колличество лайков лучшего коммента", "Название Файла",
                  "Ссылка на гифку для скачивания"])
    for dic in dic_list:
        list_of_items.append(dic["date"])
        list_of_items.append(dic["text"])
        list_of_items.append(dic['marked_as_ads'])
        list_of_items.append(dic['comments'])
        list_of_items.append(dic['text_comment'])
        list_of_items.append(dic['likes_comment'])
        list_of_items.append(filename(dic['video']))
        list_of_items.append('=HYPERLINK("{}", "{}")'.format(
            dic["video"], "Ссылка на Гифку"))
        #
        sheet.append(list_of_items)
        list_of_items = []

    book.save("VK_list.xlsx")
    print("file was done")


def combinator(a, list_, *args):
    combinator_list = []

    for i in list_:
        temp = [a] + [i] + list(args)
        temp = (tuple(temp))
        combinator_list.append(temp)

    return combinator_list


def filename(url):
    # dir = os.path.abspath(os.curdir)
    if url[-3:] == "jpg":
        file_name = url.split("/")[-1]
        file_name = file_name[:-3] + "gif"

    else:
        file_name = url.split("/")[-1][0:21] + ".gif"

    return file_name



def download(url):
    dir = os.path.abspath(os.curdir)
    if url[-3:] == "jpg":
        file_name = url.split("/")[-1]
        file_name = file_name[:-3] + "gif"

    else:
        file_name = url.split("/")[-1][0:21] + ".gif"

    print("Downloading...   ", dir + "\Output_Gif\\" + file_name)

    try:
        # open in binary mode
        with open(dir + "\Output_Gif\\" + file_name, "wb") as file:
            # get request
            response = get(url)
            # write to file
            file.write(response.content)
            file.close()
    except Exception as e:
        print((e))
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + "  ...parser...  " + str(e))
        file.close()


def mdir():
    mypath = "Output_Gif"
    if not os.path.isdir(mypath):
        os.makedirs(mypath)


def vk_parser():
    #
    #

    mdir()

    data = ""
    data1=""
    path = 'params.txt'
    with open(path, 'r') as f:
        for lines in f:
            data = data + lines

    with open('log.txt', "a") as file:
        file.write(str(datetime.datetime.now()) + data + "\n")
    d = (literal_eval(data))

    group_id_list = d["group_id_list"]
    date_start = d["date_start"]
    date_end = d["date_end"]
    comments = d["comments"]
    likes = d["likes"]
    ads = d["ads"]
    token = d["token_parser"]
    list_of = []

    # print(token)
    # token = "c245d97cc245d97cc245d97cddc228922acc245c245d97c9fd9c6af99ad638667c086c5"
    session = vk.Session(access_token=token)  # Авторизация
    vk_api = vk.API(session)

    params = combinator(vk_api, group_id_list, date_start, date_end, comments, likes, ads)

    # with Pool(10) as p:
    #     p.map(get_group_posts,params)
    #
    # print(list_of)
    #
    for i in range(len(params)):
        list_of = list_of + get_group_posts(params[i])

    excel_creator(list_of)

    for i in list_of:
        download(i['video'])


# if __name__ == "__main__":
#     # multiprocessing.freeze_support()
#     start = datetime.datetime.now()
#     main()
#     end = datetime.datetime.now()
#     print("Duration: {}".format(end - start))



def take_time(token, owner_id):
    try:
        session = vk.AuthSession(access_token=token,captcha_handler=captcha_handler)
        vk_api = vk.API(session)
        posts = vk_api.wall.get(owner_id=owner_id * (-1),
                                count=1000, filter="postponed", v=5.101)
        posts = posts["items"][-1]

        time_ = (posts["date"])
        # print("Last post finded, the time last post is {}".format(str(datetime.datetime.fromtimestamp(time_))))

        with open('log.txt', "a") as file:

            file.write("{}    Time last post is {}\n".format(str(datetime.datetime.now()),
                                                             str(datetime.datetime.fromtimestamp(time_))))

        start_shifttime = int(time_)

        return start_shifttime

    except Exception as e:
            # "list index out of range" in str(e):

            x = datetime.datetime.now()
            start_shifttime = datetime.datetime.timestamp(x)
            print("No posts, Get present time ...",
                  str(datetime.datetime.fromtimestamp(start_shifttime)))
            with open('log.txt', "a") as file:
                file.write(str(datetime.datetime.now()) + "    Time difinition error " + "   " + str(e) + "\n")
                file.write("{}    presen time is {}\n".format(str(datetime.datetime.now()),

                                                              str(datetime.datetime.fromtimestamp(start_shifttime))))

                # print(start_shifttime)
            return start_shifttime
        # else:
        #     return start_shifttime


def get_posts_number(token, owner_id):
    dic = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0, "5": 0}
    session = vk.AuthSession(access_token=token)
    vk_api = vk.API(session)
    posts = vk_api.wall.get(owner_id=owner_id * (-1),
                            count=1000, filter="postponed", v=5.101)

    return int((posts["count"]))


def poster(token, owner_id, path_folder, time_of_post):
    file_name = list_files_folder(path_folder)[0]
    #print("result")
    message = take_name_from_excel("VK_list.xlsx", file_name)

    pic_path = path_folder + file_name

    session = vk.AuthSession(access_token=token)
    vk_api = vk.API(session)

    result = vk_api.docs.getWallUploadServer(v=5.102)['upload_url']

    img = {'file': (pic_path, open(r'' + pic_path, 'rb'))}

    upload_url = result

    response = requests.post(upload_url, files=img)

    result = json.loads(response.text)
    #CAPTCHA
    #########################################################
    ## captcha_sid=788746128668,captcha_key="vaav",
    result = vk_api.docs.save(file=result['file'], v=5.102)
    #result = vk_api.docs.save(file=result['file'], captcha_sid=528554847243,captcha_key="z7kz", v=5.102)
    #########################################################

    attachments = ("doc" + str(result['doc']["owner_id"]) + "_" + str(result['doc']["id"]))

    upload = vk_api.wall.post(message=message, publish_date=time_of_post, owner_id=-1 * (owner_id), from_group=1,
                              attachments=attachments, v=5.102)

    response.close()
    return (upload, pic_path)


def timer(x):
    # x = datetime.datetime.now()
    td = int(str(x).split("-")[2].split(" ")[1].split(":")[0])
    td = 24 - td
    return td * 60 * 60


# def mdir():
#     mypath = "Uploaded_Pictures"
#     if not os.path.isdir(mypath):
#         os.makedirs(mypath)


def take_name_from_excel(file, gif_name):
    wb = load_workbook(file)

    ws = wb.active
    first_column = ws['G']
    for index, i in enumerate(first_column):
        if i.value == gif_name:
            break
    name = ws['E'][index].value
    # name = str(name)
    return name

    # print(first_column)


def list_files_folder(path):
    lst = os.listdir(path)
    lst_new = []
    if len(lst) > 0:
        for i in lst:
            if i[-3:] == "gif":
                lst_new.append(i)
        return lst_new
    else:
        return lst_new


def vk_post():
    with open('log.txt', "w") as file:
        pass
    try:
        data = ""
        data1=""
        path = 'params.txt'
        with open(path, 'r') as f:
            for lines in f:
                data = data + lines

        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + data + "\n")
        d = (literal_eval(data))



        path_folder = "Output_Gif\\"
        destination_folder = "Uploaded_Pictures\\"
        min_shift = d["min_shift"]
        token = d["token_post"]
        owner_id = d["owner_id"]
        token_parser = d["token_parser"]

    except Exception as e:
        print("read from params.txt", e)
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + "    Error on reads param.txt " + "   " + str(e) + "\n")
        file.close()

    time_zero = take_time(token, owner_id)

    uploaded_files = []

    if time_zero == None:
        print("\nServer dont response")


    elif len(list_files_folder(path_folder)) < 1:

        print("\nOutput pictures folder is empty")

    else:

        print("\nStart posting gifs")

        time_of_post = time_zero + (min_shift * 60)

        i = 1
        # poster_result = poster(token, owner_id, path_folder, time_of_post)
        # number_of_posts = get_posts_number(token, owner_id)

        while len(list_files_folder(path_folder)) > 0:

            try:
                poster_result = poster(token, owner_id, path_folder, time_of_post)
                number_of_posts = get_posts_number(token, owner_id)

                print("\npost uploaded {}, number of shft posts {}\n".format(i, number_of_posts), poster_result,
                      datetime.datetime.fromtimestamp(take_time(token, owner_id)))
                if int(poster_result[0]["post_id"]) > 0:
                    with open(poster_result[1]) as existing_file:
                        existing_file.close()
                        os.remove(poster_result[1])
                        print(" temp picture removed")

                time_of_post = time_of_post + (min_shift * 60)

                time.sleep(random.randint(12, 17) * 60)
                i = i + 1




            except Exception as error:
                #########################################################
                # print(error)
                # print(error.captcha_img)
                # print(error.captcha_sid)
                #########################################################

                error=str(error)

                if "14. Captcha needed." in error:


                    print(str(datetime.datetime.now()).split(" ")[1].split(".")[0],
                          "\n14. Error.Captcha needed, waiting....")
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + "    14. Error.Captcha needed, waiting...." + "\n")
                    time.sleep(3600)

                elif "214. Access to adding post denied: can only schedule 25 posts on a day" in error:
                    #print("ok")
                    print("\n...Оnly schedule 25 posts on a day...waiting")


                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + "214. Оnly schedule 25 posts on a day" + "\n")
                    #

                    catcha25_timer = timer(datetime.datetime.fromtimestamp(take_time(token, owner_id)))
                    time_of_post = time_of_post + catcha25_timer
                    time.sleep(random.randint(12, 17) * 60)

                elif "Access to adding post denied: cannot schedule more than 150 posts" in error:
                    # print("ok")
                    print("\n...Limit is 150 posts ...waiting")


                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + "2Limit is 150 posts ...waiting" + "\n")
                    #

                    # catcha25_timer = timer(datetime.datetime.fromtimestamp(take_time(token, owner_id)))
                    # time_of_post = time_of_post + catcha25_timer
                    time.sleep(43200)



                elif 'charmap' in error:
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + str(error) + "\n")

                    file_name_del = list_files_folder(path_folder)[0]
                    with open(file_name_del) as existing_file:
                        existing_file.close()
                        os.remove(file_name_del)


                else:
                    error = str(error)
                    print(error)
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + error + "\n")
                    break

                # time_of_post = time_of_post + (min_shift * 60)
                # print(datetime.datetime.fromtimestamp(time_of_post))


try:
    a = input("Select '0' for PARSER,\nselect '1' for POST,\nselect '2' for Parser&Post:  \n")

    if a == "0":

        vk_parser()
        input("\nParsing was done")

    elif a == "1":

        vk_post()
        # input("\nPosting was done")

    else:
        vk_parser()
        vk_post()
        # input("\nRarsin&Posting was done")
except Exception as e:
    print(e)
    with open('log.txt', "a") as file:
        file.write(str(datetime.datetime.now()) + "   " + str(e) + "\n")
    file.close()