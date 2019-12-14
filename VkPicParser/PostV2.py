import vk
import os
from ast import literal_eval
import requests
import json
import time
from openpyxl import load_workbook
import datetime
import random


def take_time(token, owner_id):
    try:
        session = vk.AuthSession(access_token=token)
        vk_api = vk.API(session)
        posts = vk_api.wall.get(owner_id=owner_id * (-1),
                                count=1000, filter="postponed", v=5.101)
        posts = posts["items"][-1]

        time_ = (posts["date"])
        # print("Last post finded, the time last post is {}".format(str(datetime.datetime.fromtimestamp(time_))))

        with open('log.txt', "a") as file:

            file.write("{}    Time last post is {}\n".format(str(datetime.datetime.now()),
                                                             str(datetime.datetime.fromtimestamp(time_))))

        start_shifttime = int(time_)

        return start_shifttime

    except Exception as e:
        if "list index out of range" in str(e):

            x = datetime.datetime.now()
            start_shifttime = datetime.datetime.timestamp(x)
            print("No posts, Get present time ...",
                  str(datetime.datetime.fromtimestamp(start_shifttime)))
            with open('log.txt', "a") as file:
                file.write(str(datetime.datetime.now()) + "    Time difinition error " + "   " + str(e) + "\n")
                file.write("{}    presen time is {}\n".format(str(datetime.datetime.now()),

                                                              str(datetime.datetime.fromtimestamp(start_shifttime))))

                # print(start_shifttime)
            return start_shifttime
        else:
            return None


def get_posts_number(token, owner_id):
    dic = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0, "5": 0}
    session = vk.AuthSession(access_token=token)
    vk_api = vk.API(session)
    posts = vk_api.wall.get(owner_id=owner_id * (-1),
                            count=1000, filter="postponed", v=5.101)

    return int((posts["count"]))


def poster(token, owner_id, path_folder, time_of_post):
    file_name = list_files_folder(path_folder)[0]

    message = take_name_from_excel("VK_list.xlsx", file_name)

    pic_path = path_folder + file_name

    session = vk.AuthSession(access_token=token)
    vk_api = vk.API(session)
    result = vk_api.docs.getWallUploadServer(v=5.102)['upload_url']

    img = {'file': (pic_path, open(r'' + pic_path, 'rb'))}

    upload_url = result

    response = requests.post(upload_url, files=img)

    result = json.loads(response.text)
    result = vk_api.docs.save(file=result['file'], v=5.102)

    attachments = ("doc" + str(result['doc']["owner_id"]) + "_" + str(result['doc']["id"]))

    upload = vk_api.wall.post(message=message, publish_date=time_of_post, owner_id=-1 * (owner_id), from_group=1,
                              attachments=attachments, v=5.102)

    response.close()
    return (upload, pic_path)


def timer(x):
    # x = datetime.datetime.now()
    td = int(str(x).split("-")[2].split(" ")[1].split(":")[0])
    td = 24 - td
    return td * 60 * 60


def mdir():
    mypath = "Uploaded_Pictures"
    if not os.path.isdir(mypath):
        os.makedirs(mypath)


def take_name_from_excel(file, gif_name):
    wb = load_workbook(file)

    ws = wb.active
    first_column = ws['G']
    for index, i in enumerate(first_column):
        if i.value == gif_name:
            break
    name = ws['E'][index].value
    # name = str(name)
    return name

    # print(first_column)


def list_files_folder(path):
    lst = os.listdir(path)
    lst_new = []
    if len(lst) > 0:
        for i in lst:
            if i[-3:] == "jpg":
                lst_new.append(i)
        return lst_new
    else:
        return lst_new


def main():
    with open('log.txt', "w") as file:
        pass
    try:
        data = ""
        path = 'params.txt'
        with open(path, 'r') as f:
            for lines in f:
                data = data + lines
        d = (literal_eval(data))

        path_folder = "Output_Pictures\\"
        destination_folder = "Uploaded_Pictures\\"
        min_shift = d["min_shift"]
        token = d["token_post"]
        owner_id = d["owner_id"]
        token_parser = d["token_parser"]

    except Exception as e:
        print("read from params.txt", e)
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + "    Error on reads param.txt " + "   " + str(e) + "\n")
        file.close()

    time_zero = take_time(token, owner_id)

    uploaded_files = []

    if time_zero == None:
        print("\nServer dont response")


    elif len(list_files_folder(path_folder)) < 1:

        print("\nOutput pictures folder is empty")

    else:

        print("\nStart posting pictures")

        time_of_post = time_zero + (min_shift * 60)
        i = 1

        while len(list_files_folder(path_folder)) > 0:

            try:

                poster_result = poster(token, owner_id, path_folder, time_of_post)
                number_of_posts = get_posts_number(token, owner_id)
                print("\npost uploaded {}, number of shft posts {}\n".format(i, number_of_posts), poster_result,
                      datetime.datetime.fromtimestamp(take_time(token, owner_id)))
                if int(poster_result[0]["post_id"]) > 0:
                    with open(poster_result[1]) as existing_file:
                        existing_file.close()
                        os.remove(poster_result[1])
                        print(" temp picture removed")

                time_of_post = time_of_post + (min_shift * 60)

                time.sleep(random.randint(8, 15) * 60)
                i = i + 1




            except Exception as error:
                # print(error)
                # print(error.captcha_img)
                # print(error.captcha_sid)
                error = str(error)
                if "14. Captcha needed." in error:
                    print(str(datetime.datetime.now()).split(" ")[1].split(".")[0], "14. Error.Captcha needed, waiting....")
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + "    14. Error.Captcha needed, waiting...." + "\n")
                    time.sleep(3600)

                elif "214. Access to adding post denied: can only schedule 25 posts on a day" in error:
                    print("Оnly schedule 25 posts on a day")
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + "214. Оnly schedule 25 posts on a day" + "\n")
                    #

                    catcha25_timer = timer(datetime.datetime.fromtimestamp(take_time(token, owner_id)))
                    time_of_post = time_of_post + catcha25_timer
                    time.sleep(random.randint(8, 15) * 60)

                else:
                    print(error)
                    with open('log.txt', "a") as file:
                        file.write(str(datetime.datetime.now()) + error + "\n")
                    break

                # time_of_post = time_of_post + (min_shift * 60)
                # print(datetime.datetime.fromtimestamp(time_of_post))


if __name__ == "__main__":
    # start = datetime.datetime.now()

    main()

    # end = datetime.datetime.now()
    # print("Duration: {}".format(end - start))
