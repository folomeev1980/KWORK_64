from openpyxl import Workbook, load_workbook
import datetime
import vk
import os
import json
import requests
from ClassesVK import Logger


import time


def photo_choice(post):
    post_dic = {}

    if post['attachments'][0]["type"] == "photo":

        photo = dict(post['attachments'][0]["photo"]["sizes"][-1])
        post_dic['video'] = photo["url"]

        return post_dic
    else:

        post_dic['video'] = None
        return post_dic


def video_choice(post):
    post_dic = {}

    if post['attachments'][0]["type"] == "photo":
        # print(post)
        photo = dict(post['attachments'][0]["photo"]["sizes"][-1])
        # post_dic['video'] = photo["url"]
        post_dic['video'] = None
        return post_dic
    else:
        # print("ok")
        # print()
        # video = dict(post['attachments'][0]["doc"]["url"])
        # print(post['attachments'][0]["doc"]["url"])
        post_dic['video'] = post['attachments'][0]["doc"]["url"]
        return post_dic


def get_comments(vk_api, owner_id, post_id):
    comments = vk_api.wall.getComments(
        owner_id=owner_id * (-1), post_id=post_id, count=100, need_likes=1, v=5.101)
    comments_number = comments["count"] // 100

    comments = (comments["items"])
    likes = comments[0]["likes"]["count"]
    text = comments[0]["text"]

    for i in range(1, comments_number + 1):
        comments = comments + vk_api.wall.getComments(
            owner_id=owner_id * (-1), post_id=post_id, need_likes=1, count=100, v=5.101, offset=i * 100)["items"]

    for index, comment in enumerate(comments):
        if comment["likes"]["count"] > likes:
            likes = comment["likes"]["count"]
            text = comment["text"]

    return {"text_comment": text, "likes_comment": likes}


def filename(url):
    # dir = os.path.abspath(os.curdir)
    if url[-3:] == "jpg":
        file_name = url.split("/")[-1]
        file_name = file_name[:-3] + "jpg"

    else:
        file_name = url.split("/")[-1][0:21] + ".gif"

    return file_name


def excel_creator(dic_list):
    print("Start to create xlx file")
    # dir = os.path.abspath(os.curdir)
    list_of_items = []

    book = Workbook()
    sheet = book.active
    # post_dic = {"date": "", "id": "", "text": "", 'marked_as_ads': "", 'video': "",
    # 'filesize': "", 'comments': "", 'text_comment': "", 'post_likes': "",
    # 'likes_comment': "", 'reposts': "", 'views': ""}
    sheet.append(["Дата и время", "Название поста", "Реклама", "Колличество комментов",
                  "Текст лучшего коммента", "Колличество лайков лучшего коммента", "Название Файла",
                  "Ссылка на гифку для скачивания"])
    for dic in dic_list:
        list_of_items.append(dic["date"])
        list_of_items.append(dic["text"])
        list_of_items.append(dic['marked_as_ads'])
        list_of_items.append(dic['comments'])
        list_of_items.append(dic['text_comment'])
        list_of_items.append(dic['likes_comment'])
        list_of_items.append(filename(dic['video']))
        list_of_items.append('=HYPERLINK("{}", "{}")'.format(
            dic["video"], "Ссылка на Объект"))
        #
        sheet.append(list_of_items)
        list_of_items = []

    book.save("VK_list.xlsx")
    print("file was done")


def take_time(token, owner_id):
    try:
        session = vk.AuthSession(access_token=token)
        vk_api = vk.API(session)
        posts = vk_api.wall.get(owner_id=owner_id * (-1),
                                count=1000, filter="postponed", v=5.101)
        posts = posts["items"][-1]

        time_ = (posts["date"])


        with open('log.txt', "a") as file:

            file.write("{}    Time last post is {}\n".format(str(datetime.datetime.now())[0:16],
                                                             str(datetime.datetime.fromtimestamp(time_))))

        start_shifttime = int(time_)

        return start_shifttime

    except Exception as e:
        x = datetime.datetime.now()
        start_shifttime = datetime.datetime.timestamp(x)
        print("No posts, Get present time ...",
              str(datetime.datetime.fromtimestamp(start_shifttime)))
        with open('log.txt', "a") as file:
            file.write(str(datetime.datetime.now()) + "    Time difinition error " + "   " + str(e) + "\n")
            file.write("{}    presen time is {}\n".format(str(datetime.datetime.now())[0:16],

                                                          str(datetime.datetime.fromtimestamp(start_shifttime))))

            # print(start_shifttime)
        return start_shifttime
    # else:
    #     return None


def get_posts_number(token, owner_id):
    session = vk.AuthSession(access_token=token)
    vk_api = vk.API(session)
    posts = vk_api.wall.get(owner_id=owner_id * (-1),
                            count=1000, filter="postponed", v=5.101)

    return int((posts["count"]))


def combinator(a, list_, *args):
    combinator_list = []

    for i in list_:
        temp = [a] + [i] + list(args)
        temp = (tuple(temp))
        combinator_list.append(temp)

    return combinator_list


def mdir(s):
    mypath = s
    if not os.path.isdir(mypath):
        os.makedirs(mypath)


def get_group_posts(*args,format_="jpg"):
    args = args[0]
    print(args)
    vk_api = args[0]
    owner_id = args[1]
    date1 = args[2]

    date2 = args[3]
    namber_comments_of = args[4]
    likes = args[5]
    ads = args[6]
    format_=format_
    post_dic = {}
    posts_list = []
    exception_numbers = 0
    date_start = datetime.datetime(date1[0], date1[1], date1[2])
    date_end = datetime.datetime(date2[0], date2[1], date2[2], 23, 59)

    posts = vk_api.wall.get(owner_id=owner_id * (-1),
                            count=100, filter=all, v=5.101)

    count_numbers = posts["count"] // 100

    posts = posts["items"]

    # Получаем первую дату поста

    for i in range(1, count_numbers + 1):
        print("Reading {} block {}".format(i, owner_id))

        temp = vk_api.wall.get(owner_id=owner_id * (-1), count=100, filter=all,
                               v=5.101, offset=i * 100)["items"]

        start_block = datetime.datetime.fromtimestamp(temp[0]["date"])
        end_block = datetime.datetime.fromtimestamp(temp[len(temp) - 1]["date"])

        if start_block > date_start:
            if end_block < date_end:
                posts = posts + temp
                print("posts block --{}-- is writting".format(i * 100))
        else:
            break

        # time.sleep(2)

    for index, post in enumerate(posts):
        try:
            d = datetime.datetime.fromtimestamp(post['date'])
            # print(d,date_start,date_end)
            if d >= date_start:
                if d <= date_end:

                    post_dic["date"] = d
                    post_dic["id"] = post["id"]
                    post_dic["text"] = post["text"]
                    post_dic['marked_as_ads'] = post['marked_as_ads']


                    # video = dict(post['attachments'][0]["doc"]["preview"]["video"])
                    if format_=="gif":
                        post_dic['video'] = (video_choice(post))['video']
                    else:
                        post_dic['video'] = (photo_choice(post))['video']

                    # post_dic['filesize'] = (dict(video)['file_size'])
                    post_dic['comments'] = post['comments']["count"]

                    if post_dic['comments'] > 0:
                        dic_ret = get_comments(vk_api, owner_id, post_dic["id"])
                        post_dic['text_comment'] = dic_ret["text_comment"]
                        post_dic['likes_comment'] = dic_ret["likes_comment"]
                    else:
                        post_dic['text_comment'] = ""
                        post_dic['likes_comment'] = ""

                    post_dic['post_likes'] = post['likes']["count"]
                    post_dic['reposts'] = post['reposts']["count"]
                    post_dic['views'] = post['views']["count"]

                    if post_dic["date"] != "":
                        if post_dic['comments'] > namber_comments_of:
                            if post_dic['likes_comment'] > likes:
                                if post_dic['marked_as_ads'] == ads:
                                    if post_dic["video"] != None:
                                        if len(post_dic['text_comment']) > 1:
                                            posts_list.append(post_dic)



        except:
            exception_numbers += 1

        if index % 10 == 0:
            print("{} posts  {} was parsed".format(owner_id, index))
        post_dic = {}
    print("Exception numbers is {} from {} pages x 100 posts ".format(
        exception_numbers, count_numbers))

    return posts_list


def download(url, direction):
    dir = os.path.abspath(os.curdir)
    if url[-3:] == "jpg":
        file_name = url.split("/")[-1]
        file_name = file_name[:-3] + "jpg"

    else:
        file_name = url.split("/")[-1][0:21] + ".gif"

    print("Downloading...   ", dir + "\{}\\".format(direction) + file_name)

    try:
        # open in binary mode
        with open(dir + "\{}\\".format(direction) + file_name, "wb") as file:
            response = requests.get(url)
            file.write(response.content)
            file.close()
    except Exception as e:

        Logger(e)


def take_name_from_excel(file, gif_name):
    wb = load_workbook(file)

    ws = wb.active
    first_column = ws['G']
    for index, i in enumerate(first_column):
        # print(i.value,gif_name)
        if i.value == gif_name:
            break
    name = ws['E'][index].value
    # name = str(name)
    return name


def timer(x):
    # x = datetime.datetime.now()
    td = int(str(x).split("-")[2].split(" ")[1].split(":")[0])
    td = 24 - td
    return td * 60 * 60


def list_files_folder(path, format_):
    lst = os.listdir(path)
    lst_new = []
    if len(lst) > 0:
        for i in lst:
            if i[-3:] == format_:
                lst_new.append(i)
        return lst_new
    else:
        return lst_new


def poster(token, owner_id, path_folder, time_of_post, format_, captcha_sid_, captcha_key_):
    file_name = list_files_folder(path_folder, format_)[0]

    message = take_name_from_excel("VK_list.xlsx", file_name)
    message = message.strip()
    # print(message,file_name)

    pic_path = path_folder + file_name

    try:

        session = vk.AuthSession(access_token=token)
        vk_api = vk.API(session)
        result = vk_api.docs.getWallUploadServer(v=5.102)['upload_url']

        #img = {'file': (pic_path, open(r'' + pic_path, 'rb'))}

        upload_url = result
        with open(r'' + pic_path, 'rb') as f:
            img = {'file': (pic_path, f)}
            with requests.post(upload_url, files=img) as r:
            #response = requests.post(upload_url, files=img)
                result = json.loads(r.text)



        # with open(r'' + pic_path, 'rb') as f:
        #     f.close()

        # CAPTCHA
        #########################################################
        ## captcha_sid=788746128668,captcha_key="vaav",
        # result = vk_api.docs.save(file=result['file'], v=5.102)
        result = vk_api.docs.save(file=result['file'], captcha_sid=captcha_sid_, captcha_key=captcha_key_, v=5.102)
        #########################################################

        attachments = ("doc" + str(result['doc']["owner_id"]) + "_" + str(result['doc']["id"]))

        upload = vk_api.wall.post(message=message, publish_date=time_of_post, owner_id=-1 * (owner_id), from_group=1,
                                  attachments=attachments, v=5.102)
        #time.sleep(30)

        # with open(pic_path) as existing_file:
        #     existing_file.close()
        #     os.remove(pic_path)
        #print("Finaly exceptin happened")

        return (upload, pic_path, message)
    except Exception as error:
        return (None, pic_path, error)
    # finally:
    #
    #     time.sleep(30)
    #
    #     r.close()
    #     with open(pic_path) as existing_file:
    #         existing_file.close()
    #     #     os.remove(pic_path)
    #     #print("Finaly exceptin happened\n"+str(error))
