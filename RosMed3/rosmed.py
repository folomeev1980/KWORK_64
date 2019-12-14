from progress.bar import Bar
import os
from selenium import webdriver
import requests
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import datetime
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import time
#########################################################################


def blanker(s):
    n = ""
    for i in s:
        if i in "\'\"<>:; .,-":
            pass
        else:
            n = n + i
    return n


def create_csv():
    try:

        with open("doctors.csv", "w", newline='') as f:
            pass
    except Exception as e:
        pass
        # print(e)


def writer_csv(data):
    try:
        for dic in data:
            for j in dic["doctors_list"]:
                with open("doctors.csv", "a", newline='') as f:
                    writer = csv.writer(f)
                    a = "{} {} {}".format(dic["famaly"], dic["name"], dic["fathername"])
                    b = "{}_{}_{}".format(j[0], j[1], j[2])
                    c = j[-1]

                    writer.writerow((b, a, c))
    except Exception as e:
        pass
        # print(e)


def read_csv():
    # dic = {}
    dic = []
    with open('doctors.csv') as csvfile:
        readCSV = csv.reader(csvfile, delimiter=',')
        for row in readCSV:
            # dic[row[0]] = row[1]
            dic.append([row[0], row[1], row[2]])

    return dic


#########################################################################

def get_html_(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'}
    r = requests.get(url, headers=headers)
    return r.text


def get_list_of_doctors(url):
    list_of_doctors = []

    if url != "":

        clinics = get_html(url)

        soup = BeautifulSoup(clinics, "lxml")
        try:
            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            # print(tds)
            for td in tds:
                temp = td.find_all("td")
                #print(temp)
                a = (temp[-2].text.strip().split("-"))

                b = (temp[-1].text.strip().split()[1][1:])

                c = temp[2]
                c = str(c).split(">")[1]
                c = c.split("<")[0].strip()

                list_of_doctors.append([a[0], a[1], b, c])
                #print(list_of_doctors)


        except Exception as e:
            print(e, "except by list of doctors")
            list_of_doctors = []

        return list_of_doctors

    else:

        return list_of_doctors


def get_append_list_of_doctors(url):
    list_of_doctors = []
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_argument('window-size=1920x935')
        driver = webdriver.Chrome('chromedriver.exe', options=options)
        driver.get(url)
    except Exception as e:

        driver.quit()

    k = 2

    while True:

        try:
            # print("ok")
            page = "/html/body/form/table/tbody/tr/td/table/tbody/tr[3]/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr[22]/td/table/tbody/tr/td[{}]/a".format(
                k)
            driver.find_element_by_xpath(page).click()
            html = driver.page_source
            # print(html)

            soup = BeautifulSoup(html, "lxml")

            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                c = temp[2]
                c = str(c).split(">")[1]
                c = c.split("<")[0].strip()

                a = (temp[-2].text.strip().split("-"))

                b = (temp[-1].text.strip().split()[1][1:])
                list_of_doctors.append([a[0], a[1], b, c])

            k = k + 1
            #time.sleep(2)


        except Exception as e:
            # print("End pages",e)
            break
            # driver.quit()
    driver.quit()

    return list_of_doctors


def get_info_for_each_page_(html):
    list_dics = []
    soup = BeautifulSoup(html, "lxml")

    list_of_page = (soup.find_all("tr", class_="hi_sys poi"))

    for st, page in enumerate(list_of_page):

        dic = {}
        list_some = []

        try:
            doctor_link = "https://grls.rosminzdrav.ru/" + \
                          page.get("onclick").split("'")[1]

            doctors_list = get_list_of_doctors(doctor_link)



        except:
            doctors_list = []

        try:

            page_data = page.find_all("td")

            for td in page_data:
                list_some.append(td.text.strip())

            # print(list_some[6])

            if int(list_some[6]) <= 20:
                list_some.append(doctors_list)
            else:
                list_some.append(doctors_list)
                # print(doctor_link)
                list_some[-1].extend(get_append_list_of_doctors(doctor_link))

            # print(list_some)

            dic["famaly"] = list_some[2]
            dic["name"] = list_some[3]
            dic["fathername"] = list_some[4]
            dic["numer_of_ki"] = list_some[6]
            dic["doctors_list"] = list_some[8]
            # print(dic["doctors_list"])

            list_dics.append(dic)


        except Exception as e:
            print(e, "Ошибка ")
            dic = {}
            list_dics.append(dic)



    return list_dics


def cvs_doctors():
    i = 1
    k = 0
    sum = 0
    create_csv()
    while True:
        try:
            url = "https://grls.rosminzdrav.ru/ciexperts.aspx?F=&N=&P=&D=&ExpertID=&order=fio&orderType=desc&moduleId=2&pageSize=30&pageNum={}".format(
                i)
            html = get_html_(url)
            list_of_dics_one_page = get_info_for_each_page_(html)
            if len(list_of_dics_one_page) == 0:
                k = k + 1
            sum = sum + len(list_of_dics_one_page)
            writer_csv(list_of_dics_one_page)
            i = i + 1
            print(sum)
            if k > 5:
                break
        except Exception as e:
            if k > 5:
                break
    print(i, "Pages\n", sum, "Doctors")


#########################################################################

def func(list_):
    s = ""
    for index, i in enumerate(list_):
        s = s + str(i) + "\n"
    return s


def excel_append_row(n):
    book = load_workbook("Med_list.xlsx")
    print("ok")
    sheet = book.worksheets[0]
    sheet.insert_rows(idx=2, amount=n)
    book.save("Med_list_p2.xlsx")


def excel_check():
    if os.path.exists("Med_list.xlsx"):
        print("Файл Med_list.xlsx найден")
        book = load_workbook("Med_list.xlsx", read_only=True)
        sheet = book["Sheet"]

        rki = sheet.cell(row=2, column=2)
        date_rki = sheet.cell(row=2, column=3)

        # print(rki,date_rki)
        return (str(rki.value), str(date_rki.value))


    else:
        print("Файл не найден")
        return False


def get_html(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'}
    r = requests.get(url, headers=headers)
    return r.text


def get_info_for_each_page(html):
    list_dics = []
    soup = BeautifulSoup(html, "lxml")
    list_of_page = (soup.find_all("tr", class_="hi_sys poi stat_reged"))

    for st, page in enumerate(list_of_page):

        dic = {}
        list_some = []
        try:
            url_clinic = "https://grls.rosminzdrav.ru/" + \
                         page.get("onclick").split("'")[1]

        except:
            url_clinic = ""

        clinc_lists = get_list_of_clinics(url_clinic)

        try:
            page_data = page.find_all("td")

            for td in page_data:
                list_some.append(td.text.strip())
            list_some.append(clinc_lists)

            dic["n_n"] = list_some[0]
            dic["rki"] = list_some[1]
            dic["date"] = list_some[2]
            dic["name_lp"] = list_some[3]
            dic["organization"] = list_some[4]
            dic["country"] = list_some[5]
            dic["organization_lp"] = list_some[6]
            dic["date_start"] = list_some[7]
            dic["date_end"] = list_some[8]
            dic["n_protokol"] = list_some[9]
            dic["protokol"] = list_some[10]
            dic["phase"] = list_some[11]
            dic["view"] = list_some[12]
            dic["n_orgs"] = list_some[13]
            dic["n_patient"] = list_some[14]
            dic["type"] = list_some[15]
            dic["status"] = list_some[16]
            dic["clinc_listsl"] = list_some[17]
            dic["cells"] = (len(clinc_lists))
            list_dics.append(dic)
        except:
            dic = {}
            list_dics.append(dic)
        print(st, dic["rki"], dic["date"])
        # time.sleep(random.randint(0,3))

    return list_dics


def excel_creator(dic_list, new=True):
    row_enumerate = 2
    print("Start to create xlx file")

    list_of_items = []
    if new:
        book = Workbook()
        sheet = book.active
    else:
        # dic_list=dic_list[:-1]
        book1 = load_workbook("Med_list.xlsx", read_only=True)
        sheet = book1["Sheet"]

        book = Workbook()
        sheet_book = book.active

        for inx, row in enumerate(sheet.rows):
            for cl, t in enumerate(row):
                sheet_book.cell(row=inx + 1, column=cl + 1).value = t.value
            if sheet_book.cell(row=inx + 1, column=cl + 1).value == 1:
                sheet_book.row_dimensions[inx + 1].height = 60

        book1.close()

        try:
            shl = (book.sheetnames)
            if len(shl) > 0:
                for h in shl:
                    if h != "Sheet":
                        book.remove(book[h])
        except:
            pass

        book.create_sheet("Sheet_row")
        sheet = book["Sheet_row"]

    if new:
        sheet.append(["номер п/п",
                      "Номер РКИ",
                      "Дата создания РКИ",
                      "Наименование ЛП",
                      "Организация, проводящая КИ",
                      "Страна разраб-ка",
                      "Организация, привлеченная разработчиком ЛП",
                      "Начало (дата)",
                      "Окончание (дата)",
                      "№ протокола",
                      "Протокол",
                      "Фаза КИ",
                      "Вид КИ",
                      "Колич. мед. орг-й",
                      "Колич. пациент.",
                      "Области применения",
                      "Состояние",
                      "Перечень медицинских организаций, в которых предполагается проведение клинических исследований"])

    bar = Bar('Excel_writer', max=len(dic_list))
    for index, d in enumerate(dic_list):
        for i in range(d["cells"]):
            step = d["cells"]
            list_of_items.append((d["n_n"]))
            list_of_items.append(d["rki"])
            list_of_items.append(d["date"])
            list_of_items.append(d["name_lp"])
            list_of_items.append(d["organization"])
            list_of_items.append(d["country"])
            list_of_items.append(d["organization_lp"])
            list_of_items.append(d["date_start"])
            list_of_items.append(d["date_end"])
            list_of_items.append(d["n_protokol"])
            list_of_items.append(d["protokol"])
            list_of_items.append(d["phase"])
            list_of_items.append(d["view"])
            list_of_items.append(d["n_orgs"])
            list_of_items.append(d["n_patient"])
            list_of_items.append(d["type"])
            list_of_items.append(d["status"])
            # list_of_items.append(func(d["clinc_listsl"]))
            list_of_items.append((d["clinc_listsl"][i]))
            list_of_items.append(d["cells"])
            if d["rki"] != "":
                sheet.append(list_of_items)

            list_of_items = []

        bar.next()

    bar.finish()

    r = len(sheet['A'])
    # print("Create Sheet_row")

    if new == False:
        sheet_d = book["Sheet_row"]
        sheet = book["Sheet"]
        sheet.insert_rows(idx=2, amount=r)

        for ind, row in enumerate(sheet_d.rows):
            for col, k in enumerate(row):
                sheet.cell(row=ind + 2, column=col + 1).value = k.value

    book.save("Med_list.xlsx")

    print("file was done")


def get_list_of_clinics(url):
    list_of_clinics = []

    if url != "":

        clinics = get_html(url)

        soup = BeautifulSoup(clinics, "lxml")
        try:
            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                for j in temp:
                    if len(j.text) > 5:
                        list_of_clinics.append(j.text.strip())
        except Exception as e:
            list_of_clinics = []

        return list_of_clinics

    else:

        return list_of_clinics


def excel_edit(dic):
    book1 = load_workbook("Med_list.xlsx", read_only=True)
    sheet = book1["Sheet"]

    book_new = Workbook()
    sheet_book = book_new.active

    for inx, row in enumerate(sheet.rows):
        for cl, t in enumerate(row):
            sheet_book.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_book.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_book.row_dimensions[inx + 1].height = 60

    book1.close()

    book_new.create_sheet("KI_Edit")

    # sheet = book_new["Sheet"]
    sheet_new = book_new["KI_Edit"]

    for inx, row in enumerate(sheet_book.rows):
        for cl, t in enumerate(row):
            sheet_new.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_new.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_new.row_dimensions[inx + 1].height = 60

    sheet_new.insert_cols(19)
    sheet_new.insert_cols(20)

    reg = ["Онколог", "онколог", "Патолог", "патолог"]

    bar = Bar('Processing', max=len(sheet_new['A']))

    for inx, row in enumerate(sheet_new.rows):

        s = "{}_{}_{}".format(sheet_new.cell(row=inx + 1, column=8).value,
                              sheet_new.cell(row=inx + 1, column=9).value,
                              sheet_new.cell(row=inx + 1, column=2).value)
        # print(s)
        try:
            for i in dic:
                # print(blanker(i[2]))
                # print(blanker(str(sheet_new.cell(row=inx + 1, column=18).value)))
                if s == i[0] and blanker(i[2]) in blanker(str(sheet_new.cell(row=inx + 1, column=18).value)):
                    # print("ok")
                    sheet_new.cell(row=inx + 1, column=19).value = i[1]
                    break
        except Exception:
            sheet_new.cell(row=inx + 1, column=19).value = ""

        for j in reg:
            if j in str(sheet_new.cell(row=inx + 1, column=11).value) + " " + str \
                        (sheet_new.cell(row=inx + 1, column=16).value):
                sheet_new.cell(row=inx + 1, column=20).value = pacient(
                    str(sheet_new.cell(row=inx + 1, column=11).value))
        bar.next()

    bar.finish()

    # book_new.save("test.xlsx")

    #####

    print("Alignment 1 iter KI")
    bar = Bar('Processing', max=len(sheet_new['A']))

    checker = int(sheet_new.cell(row=2, column=21).value)
    checker2 = int(sheet_new.cell(row=2, column=2).value)
    print((checker, checker2))

    for liner, i in enumerate(range(2, len(sheet_new['A']) + 1)):

        if int(sheet_new.cell(row=i, column=21).value) == checker and int(
                sheet_new.cell(row=i, column=2).value) == checker2:
            # print(i, checker)
            # print("pass")
            pass
        else:
            # print(liner)
            for col in range(1, 18):
                # print(i, checker, col)
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            for col in range(20, 21):
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            checker = int(sheet_new.cell(row=i, column=21).value)
            checker2 = int(sheet_new.cell(row=i, column=2).value)

        bar.next()

    bar.finish()

    for col in range(1, 18):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)
    for col in range(20, 21):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)

    print("Alignment2 iter KI")

    bar = Bar('Processing', max=len(sheet_new['A']))

    for row in sheet_new.rows:
        for index, k in enumerate(row):
            if index != 17:
                k.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
                k.font = Font(size="9")
            else:
                k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
                k.font = Font(size="9")
        #
        bar.next()
    bar.finish()

    #
    # print("Alignment 3 iter KI")
    # bar = Bar('Processing', max=len(sheet_new['A']))
    # for row in sheet_new.rows:
    #     for k in row[-2:]:
    #         k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
    #
    #     bar.next()
    # bar.finish()

    sheet_new.column_dimensions[get_column_letter(4)].width = 25
    sheet_new.column_dimensions[get_column_letter(5)].width = 35
    sheet_new.column_dimensions[get_column_letter(7)].width = 50
    sheet_new.column_dimensions[get_column_letter(11)].width = 60
    sheet_new.column_dimensions[get_column_letter(16)].width = 11
    sheet_new.column_dimensions[get_column_letter(17)].width = 11
    sheet_new.column_dimensions[get_column_letter(18)].width = 200
    sheet_new.column_dimensions[get_column_letter(19)].width = 30
    sheet_new.column_dimensions[get_column_letter(20)].width = 60

    # book.close()
    book_new.save("Med_list.xlsx")


def pacient(s):
    l = ["пациент", "Пациент", "Детей", "детей"]

    for i in l:
        ind = s.find(i)
        if ind != -1:
            return "Для " + s[ind:]
        else:
            return ""


def main():

    def creator():
        flag = True

        full_list_of_dic = []
        k = 1
        lenght_tables = 10000

        url_orig = "https://grls.rosminzdrav.ru/CiPermitionReg.aspx?PermYear=0&DateBeg=&DateEnd=&DateInc=&NumInc=&RegNm=&Statement=&Protocol=&Qualifier=&ProtoNum=&idCIStatementCh=&CiPhase=&RangeOfApp=&Torg=&LFDos=&Producer=&Recearcher=&sponsorCountry=&MedBaseCount=&CiType=&PatientCount=&OrgDocOut=2&Status=1&NotInReg=0&All=0&PageSize={}&order=date_perm&orderType=desc&pagenum=".format(
            lenght_tables)

        try:

            while k <= 1:
                url = url_orig + str(k)
                html = get_html(url)
                list_of_dics = get_info_for_each_page(html)
                full_list_of_dic.extend(list_of_dics)
                k = k + 1


        except Exception as e:

            print(e, k)

        print("End reading IK informatiom")

        excel_creator(full_list_of_dic, True)

    i = input(
        "Chose action:\n1 - parse db researches\n2 - parse db doctors\n3 - generate xls file\n4 - for all actions\n\n")

    if i == "1":
        creator()

    elif i == "2":

        cvs_doctors()
    elif i == "3":

        dic = read_csv()
        excel_edit(dic)

    else:

        creator()
        cvs_doctors()
        excel_edit(read_csv())


if __name__ == "__main__":
    start = datetime.datetime.now()
    main()
    end = datetime.datetime.now()
    print("Duration: {}".format(end - start))
    input("\nPRESS any batton to complete.....")
