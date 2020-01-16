from progress.bar import Bar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import csv







def blanker(s):
    n = ""
    for i in s:
        if i in "\'\"<>:; .,-":
            pass
        else:
            n = n + i
    return n




def read_csv():
    # dic = {}
    dic = []
    with open('doctors.csv') as csvfile:
        readCSV = csv.reader(csvfile, delimiter=',')
        for row in readCSV:
            # dic[row[0]] = row[1]
            dic.append([row[0], row[1], row[2]])

    return dic













def excel_edit(dic):
    book1 = load_workbook("Med_list.xlsx", read_only=True)
    sheet = book1["Sheet"]

    book_new = Workbook()
    sheet_book = book_new.active

    for inx, row in enumerate(sheet.rows):
        for cl, t in enumerate(row):
            sheet_book.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_book.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_book.row_dimensions[inx + 1].height = 60

    book1.close()

    book_new.create_sheet("KI_Edit")

    # sheet = book_new["Sheet"]
    sheet_new = book_new["KI_Edit"]

    for inx, row in enumerate(sheet_book.rows):
        for cl, t in enumerate(row):
            sheet_new.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_new.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_new.row_dimensions[inx + 1].height = 60

    sheet_new.insert_cols(19)
    sheet_new.insert_cols(20)

    reg = ["Онколог", "онколог", "Патолог", "патолог"]

    bar = Bar('Processing', max=len(sheet_new['A']))

    for inx, row in enumerate(sheet_new.rows):

        s = "{}_{}_{}".format(sheet_new.cell(row=inx + 1, column=8).value,
                              sheet_new.cell(row=inx + 1, column=9).value,
                              sheet_new.cell(row=inx + 1, column=2).value)
        # print(s)
        try:
            for i in dic:
                # print(blanker(i[2]))
                # print(blanker(str(sheet_new.cell(row=inx + 1, column=18).value)))
                if s == i[0] and blanker(i[2]) in blanker(str(sheet_new.cell(row=inx + 1, column=18).value)):
                    # print("ok")
                    sheet_new.cell(row=inx + 1, column=19).value = i[1]
                    break
        except Exception:
            sheet_new.cell(row=inx + 1, column=19).value = ""

        for j in reg:
            if j in str(sheet_new.cell(row=inx + 1, column=11).value) + " " + str \
                        (sheet_new.cell(row=inx + 1, column=16).value):
                sheet_new.cell(row=inx + 1, column=20).value = pacient(
                    str(sheet_new.cell(row=inx + 1, column=11).value))
        bar.next()

    bar.finish()

    # book_new.save("test.xlsx")

    #####

    print("Alignment 1 iter KI")
    bar = Bar('Processing', max=len(sheet_new['A']))

    checker = int(sheet_new.cell(row=2, column=21).value)
    checker2 = int(sheet_new.cell(row=2, column=2).value)
    print((checker, checker2))

    for liner, i in enumerate(range(2, len(sheet_new['A']) + 1)):

        if int(sheet_new.cell(row=i, column=21).value) == checker and int(
                sheet_new.cell(row=i, column=2).value) == checker2:
            # print(i, checker)
            # print("pass")
            pass
        else:
            # print(liner)
            for col in range(1, 18):
                # print(i, checker, col)
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            for col in range(20, 21):
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            checker = int(sheet_new.cell(row=i, column=21).value)
            checker2 = int(sheet_new.cell(row=i, column=2).value)

        bar.next()

    bar.finish()

    for col in range(1, 18):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)
    for col in range(20, 21):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)

    print("Alignment2 iter KI")

    bar = Bar('Processing', max=len(sheet_new['A']))

    for row in sheet_new.rows:
        for index, k in enumerate(row):
            if index != 17:
                k.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
                k.font = Font(size="9")
            else:
                k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
                k.font = Font(size="9")
        #
        bar.next()
    bar.finish()



    sheet_new.column_dimensions[get_column_letter(4)].width = 25
    sheet_new.column_dimensions[get_column_letter(5)].width = 35
    sheet_new.column_dimensions[get_column_letter(7)].width = 50
    sheet_new.column_dimensions[get_column_letter(11)].width = 60
    sheet_new.column_dimensions[get_column_letter(16)].width = 11
    sheet_new.column_dimensions[get_column_letter(17)].width = 11
    sheet_new.column_dimensions[get_column_letter(18)].width = 200
    sheet_new.column_dimensions[get_column_letter(19)].width = 30
    sheet_new.column_dimensions[get_column_letter(20)].width = 60

    # book.close()
    book_new.save("Med_list.xlsx")


def pacient(s):
    l = ["пациент", "Пациент", "Детей", "детей"]

    for i in l:
        ind = s.find(i)
        if ind != -1:
            return "Для " + s[ind:]
        else:
            return ""