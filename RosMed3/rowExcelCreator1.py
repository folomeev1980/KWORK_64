from progress.bar import Bar
import requests
import csv
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from random import randint
from time import sleep


def create_csv():
    header = tuple(["номер п/п",
                    "Номер РКИ",
                    "Дата создания РКИ",
                    "Наименование ЛП",
                    "Организация, проводящая КИ",
                    "Страна разраб-ка",
                    "Организация, привлеченная разработчиком ЛП",
                    "Начало (дата)",
                    "Окончание (дата)",
                    "№ протокола",
                    "Протокол",
                    "Фаза КИ",
                    "Вид КИ",
                    "Колич. мед. орг-й",
                    "Колич. пациент.",
                    "Области применения",
                    "Состояние",
                    "Перечень медицинских организаций, в которых предполагается проведение клинических исследований",
                    "Колличество медицинских организаций"])

    with open("clinics.csv", "w", newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(header)


def get_html(url):
    page = ""
    while page == '':
        try:
            headers = {
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'}
            page = requests.get(url, headers=headers)
            return page.text
            break
        except:
            print("Connection refused by the server..")
            print("Let me sleep for 10 seconds")
            # print("ZZzzzz...")
            sleep(10)
            # print("Was a nice sleep, now let me continue...")
            continue


def get_list_of_clinics(url):
    list_of_clinics = []

    if url != "":

        clinics = get_html(url)

        soup = BeautifulSoup(clinics, "lxml")
        try:
            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                for j in temp:
                    if len(j.text) > 5:
                        list_of_clinics.append(j.text.strip())
        except Exception as e:
            list_of_clinics = []

        return list_of_clinics

    else:

        return list_of_clinics


def get_info_for_each_page(html):
    list_dics = []
    soup = BeautifulSoup(html, "lxml")
    list_of_page = (soup.find_all("tr", class_="hi_sys poi stat_reged"))

    for st, page in enumerate(list_of_page):

        dic = {}
        list_some = []
        try:
            url_clinic = "https://grls.rosminzdrav.ru/" + \
                         page.get("onclick").split("'")[1]

        except:
            url_clinic = ""

        clinc_lists = get_list_of_clinics(url_clinic)

        try:
            page_data = page.find_all("td")

            for td in page_data:
                list_some.append(td.text.strip())
            list_some.append(clinc_lists)

            dic["n_n"] = list_some[0]
            dic["rki"] = list_some[1]
            dic["date"] = list_some[2]
            dic["name_lp"] = list_some[3]
            dic["organization"] = list_some[4]
            dic["country"] = list_some[5]
            dic["organization_lp"] = list_some[6]
            dic["date_start"] = list_some[7]
            dic["date_end"] = list_some[8]
            dic["n_protokol"] = list_some[9]
            dic["protokol"] = list_some[10]
            dic["phase"] = list_some[11]
            dic["view"] = list_some[12]
            dic["n_orgs"] = list_some[13]
            dic["n_patient"] = list_some[14]
            dic["type"] = list_some[15]
            dic["status"] = list_some[16]
            dic["clinc_listsl"] = list_some[17]
            dic["cells"] = (len(clinc_lists))
            list_dics.append(dic)
        except:
            dic = {}
            list_dics.append(dic)

        print(st, dic["rki"], dic["date"])
        sleep(randint(0, 1))

    return list_dics


def excel_creator(dic_list, new=True):
    row_enumerate = 2
    print("Start to create xlx file")

    list_of_items = []
    if new:
        book = Workbook()
        sheet = book.active
    else:
        # dic_list=dic_list[:-1]
        book1 = load_workbook("Med_list.xlsx", read_only=True)
        sheet = book1["Sheet"]

        book = Workbook()
        sheet_book = book.active

        for inx, row in enumerate(sheet.rows):
            for cl, t in enumerate(row):
                sheet_book.cell(row=inx + 1, column=cl + 1).value = t.value
            if sheet_book.cell(row=inx + 1, column=cl + 1).value == 1:
                sheet_book.row_dimensions[inx + 1].height = 60

        book1.close()

        try:
            shl = (book.sheetnames)
            if len(shl) > 0:
                for h in shl:
                    if h != "Sheet":
                        book.remove(book[h])
        except:
            pass

        book.create_sheet("Sheet_row")
        sheet = book["Sheet_row"]

    if new:
        sheet.append(["номер п/п",
                      "Номер РКИ",
                      "Дата создания РКИ",
                      "Наименование ЛП",
                      "Организация, проводящая КИ",
                      "Страна разраб-ка",
                      "Организация, привлеченная разработчиком ЛП",
                      "Начало (дата)",
                      "Окончание (дата)",
                      "№ протокола",
                      "Протокол",
                      "Фаза КИ",
                      "Вид КИ",
                      "Колич. мед. орг-й",
                      "Колич. пациент.",
                      "Области применения",
                      "Состояние",
                      "Перечень медицинских организаций, в которых предполагается проведение клинических исследований"])

    bar = Bar('Excel_writer', max=len(dic_list))
    for index, d in enumerate(dic_list):
        for i in range(d["cells"]):
            # step = d["cells"]
            list_of_items.append((d["n_n"]))
            list_of_items.append(d["rki"])
            list_of_items.append(d["date"])
            list_of_items.append(d["name_lp"])
            list_of_items.append(d["organization"])
            list_of_items.append(d["country"])
            list_of_items.append(d["organization_lp"])
            list_of_items.append(d["date_start"])
            list_of_items.append(d["date_end"])
            list_of_items.append(d["n_protokol"])
            list_of_items.append(d["protokol"])
            list_of_items.append(d["phase"])
            list_of_items.append(d["view"])
            list_of_items.append(d["n_orgs"])
            list_of_items.append(d["n_patient"])
            list_of_items.append(d["type"])
            list_of_items.append(d["status"])
            # list_of_items.append(func(d["clinc_listsl"]))
            list_of_items.append((d["clinc_listsl"][i]))
            list_of_items.append(d["cells"])
            if d["rki"] != "":
                sheet.append(list_of_items)

            list_of_items = []

        bar.next()

    bar.finish()

    r = len(sheet['A'])
    # print("Create Sheet_row")

    if new == False:
        sheet_d = book["Sheet_row"]
        sheet = book["Sheet"]
        sheet.insert_rows(idx=2, amount=r)

        for ind, row in enumerate(sheet_d.rows):
            for col, k in enumerate(row):
                sheet.cell(row=ind + 2, column=col + 1).value = k.value

    book.save("Med_list.xlsx")

    print("file was done")


def csv_row_clinic_creator(dic_list):
    list_of_items = []
    k = 0

    # bar = Bar('Excel_writer', max=len(dic_list))
    for d in dic_list:
        try:
            for i in range(d["cells"]):
                # step = d["cells"]
                list_of_items.append((d["n_n"]))
                list_of_items.append(d["rki"])
                list_of_items.append(d["date"])
                list_of_items.append(d["name_lp"])
                list_of_items.append(d["organization"])
                list_of_items.append(d["country"])
                list_of_items.append(d["organization_lp"])
                list_of_items.append(d["date_start"])
                list_of_items.append(d["date_end"])
                list_of_items.append(d["n_protokol"])
                list_of_items.append(d["protokol"])
                list_of_items.append(d["phase"])
                list_of_items.append(d["view"])
                list_of_items.append(d["n_orgs"])
                list_of_items.append(d["n_patient"])
                list_of_items.append(d["type"])
                list_of_items.append(d["status"])
                # list_of_items.append(func(d["clinc_listsl"]))
                list_of_items.append((d["clinc_listsl"][i]))
                list_of_items.append(d["cells"])
                if d["rki"] != "":
                    # temp=[]
                    # for some in list_of_items:
                    #     temp.append(some.decode("utf-8", "ignore"))

                    with open("clinics.csv", "a", newline='', encoding='utf-8') as f:
                        writer = csv.writer(f, delimiter=';')
                        # print(tuple(list_of_items))
                        writer.writerow(tuple(list_of_items))
                    # temp=[]

                list_of_items = []
        except Exception as e:
            print(list_of_items)
            print("Exception csv writer", k, e)
            list_of_items = []
            k += 1


def creator():
    create_csv()
    full_list_of_dic = []
    k = 1
    lenght_tables = 10000

    url_orig = "https://grls.rosminzdrav.ru/CiPermitionReg.aspx?PermYear=0&DateBeg=&DateEnd=&DateInc=&NumInc=&RegNm=&Statement=&Protocol=&Qualifier=&ProtoNum=&idCIStatementCh=&CiPhase=&RangeOfApp=&Torg=&LFDos=&Producer=&Recearcher=&sponsorCountry=&MedBaseCount=&CiType=&PatientCount=&OrgDocOut=2&Status=1&NotInReg=0&All=0&PageSize={}&order=date_perm&orderType=desc&pagenum=".format(
        lenght_tables)

    try:

        while k <= 1:
            url = url_orig + str(k)
            html = get_html(url)
            list_of_dics = get_info_for_each_page(html)
            full_list_of_dic.extend(list_of_dics)

            k = k + 1


    except Exception as e:
        print(e, k)

    print("End reading IK informatiom")
    # print(len(full_list_of_dic))
    csv_row_clinic_creator(full_list_of_dic)
    # excel_creator(full_list_of_dic, True)
