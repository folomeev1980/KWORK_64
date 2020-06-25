from progress.bar import Bar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import csv

list_of_lines = []


def csv_reader():
    with open('medlist_combination.csv', newline='') as medlist:
        csv_reader = csv.reader(medlist, delimiter=';')

        for index, line in enumerate(csv_reader):
            if index > 0:
                list_of_lines.append(line)


def excel_maker(lines):
    book_new = Workbook()
    sheet_new = book_new.active
    header = tuple(["номер п/п",
                    "Номер РКИ",
                    "Дата создания РКИ",
                    "Наименование ЛП",
                    "Организация, проводящая КИ",
                    "Страна разраб-ка",
                    "Организация, привлеченная разработчиком ЛП",
                    "Начало (дата)",
                    "Окончание (дата)",
                    "№ протокола",
                    "Протокол",
                    "Фаза КИ",
                    "Вид КИ",
                    "Колич. мед. орг-й",
                    "Колич. пациент.",
                    "Области применения",
                    "Состояние",
                    # "Перечень медицинских организаций, в которых предполагается проведение клинических исследований",
                    # "ФИО Главного Исследователя",
                    # "Код врача",
                    # "День рождения врача",
                    # "Для чего",
                    # "Критерии Включения",
                    # "Критерии Исключения"
                    ])

    sheet_new.append(header)
    for i in lines:
        sheet_new.append(i)

    checker = int(sheet_new.cell(row=2, column=14).value)
    checker2 = int(sheet_new.cell(row=2, column=2).value)
    print((checker, checker2))

    # for liner, i in enumerate(range(2, len(sheet_new['A']) + 1)):
    for liner, i in enumerate(range(2, len(sheet_new['A']) + 1)):

        if int(sheet_new.cell(row=i, column=14).value) == checker and int(
                sheet_new.cell(row=i, column=2).value) == checker2:
            # print(i, checker)
            # print("pass")
            pass
        else:
            # print(liner)
            for col in range(1, 18):
                # print(i, checker, col)
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            for col in range(22, 25):
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            checker = int(sheet_new.cell(row=i, column=14).value)
            checker2 = int(sheet_new.cell(row=i, column=2).value)

        if liner % 100 == 0:
            print(liner)

    for col in range(1, 18):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)
    for col in range(22, 25):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)

    for row in sheet_new.rows:
        for index, k in enumerate(row):
            if index == 17 or index == 22 or index == 23:
                k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
                k.font = Font(size="9")
            else:
                k.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
                k.font = Font(size="9")

    sheet_new.column_dimensions[get_column_letter(4)].width = 25
    sheet_new.column_dimensions[get_column_letter(5)].width = 35
    sheet_new.column_dimensions[get_column_letter(7)].width = 50
    sheet_new.column_dimensions[get_column_letter(11)].width = 60
    sheet_new.column_dimensions[get_column_letter(16)].width = 11
    sheet_new.column_dimensions[get_column_letter(17)].width = 11
    sheet_new.column_dimensions[get_column_letter(18)].width = 150
    sheet_new.column_dimensions[get_column_letter(19)].width = 30
    sheet_new.column_dimensions[get_column_letter(20)].width = 10
    sheet_new.column_dimensions[get_column_letter(21)].width = 10
    sheet_new.column_dimensions[get_column_letter(22)].width = 70
    sheet_new.column_dimensions[get_column_letter(23)].width = 90
    sheet_new.column_dimensions[get_column_letter(24)].width = 90

    book_new.save("RESEARCH.xlsx")


def excel():
    csv_reader()
    print(len(list_of_lines))
    excel_maker(list_of_lines)
