from progress.bar import Bar
import os
from ast import literal_eval
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import datetime
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random


def func(list_):
    s = ""
    for index, i in enumerate(list_):
        s = s + str(i) + "\n"
    return s


def excel_append_row(n):
    book = load_workbook("Med_list.xlsx")
    print("ok")
    sheet = book.worksheets[0]
    sheet.insert_rows(idx=2, amount=n)
    book.save("Med_list_p2.xlsx")


def excel_check():
    if os.path.exists("Med_list.xlsx"):
        print("Файл Med_list.xlsx найден")
        book = load_workbook("Med_list.xlsx", read_only=True)
        sheet = book["Sheet"]

        rki = sheet.cell(row=2, column=2)
        date_rki = sheet.cell(row=2, column=3)

        # print(rki,date_rki)
        return (str(rki.value), str(date_rki.value))


    else:
        print("Файл не найден")
        return False


def get_html(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'}
    r = requests.get(url, headers=headers)
    return r.text


def get_info_for_each_page(html):
    list_dics = []
    soup = BeautifulSoup(html, "lxml")
    list_of_page = (soup.find_all("tr", class_="hi_sys poi stat_reged"))

    for st, page in enumerate(list_of_page):

        dic = {}
        list_some = []
        try:
            url_clinic = "https://grls.rosminzdrav.ru/" + \
                         page.get("onclick").split("'")[1]

        except:
            url_clinic = ""

        clinc_lists = get_list_of_clinics(url_clinic)

        try:
            page_data = page.find_all("td")

            for td in page_data:
                list_some.append(td.text.strip())
            list_some.append(clinc_lists)

            dic["n_n"] = list_some[0]
            dic["rki"] = list_some[1]
            dic["date"] = list_some[2]
            dic["name_lp"] = list_some[3]
            dic["organization"] = list_some[4]
            dic["country"] = list_some[5]
            dic["organization_lp"] = list_some[6]
            dic["date_start"] = list_some[7]
            dic["date_end"] = list_some[8]
            dic["n_protokol"] = list_some[9]
            dic["protokol"] = list_some[10]
            dic["phase"] = list_some[11]
            dic["view"] = list_some[12]
            dic["n_orgs"] = list_some[13]
            dic["n_patient"] = list_some[14]
            dic["type"] = list_some[15]
            dic["status"] = list_some[16]
            dic["clinc_listsl"] = list_some[17]
            dic["cells"] = (len(clinc_lists))
            list_dics.append(dic)
        except:
            dic = {}
            list_dics.append(dic)
        print(st, dic["rki"], dic["date"])
        # time.sleep(random.randint(0,3))

    return list_dics


def excel_creator(dic_list, new=True):
    row_enumerate = 2
    print("Start to create xlx file")

    list_of_items = []
    if new:
        book = Workbook()
        sheet = book.active
    else:
        # dic_list=dic_list[:-1]
        book = load_workbook("Med_list.xlsx")
        try:
            shl = (book.sheetnames)
            if len(shl) > 0:
                for h in shl:
                    if h != "Sheet":
                        book.remove(book[h])
        except:
            pass

        book.create_sheet("Sheet_row")
        sheet = book["Sheet_row"]

    if new:
        sheet.append(["номер п/п",
                      "Номер РКИ",
                      "Дата создания РКИ",
                      "Наименование ЛП",
                      "Организация, проводящая КИ",
                      "Страна разраб-ка",
                      "Организация, привлеченная разработчиком ЛП",
                      "Начало (дата)",
                      "Окончание (дата)",
                      "№ протокола",
                      "Протокол",
                      "Фаза КИ",
                      "Вид КИ",
                      "Колич. мед. орг-й",
                      "Колич. пациент.",
                      "Области применения",
                      "Состояние",
                      "Перечень медицинских организаций, в которых предполагается проведение клинических исследований"])

    for index, d in enumerate(dic_list):
        for i in range(d["cells"]):
            step = d["cells"]
            list_of_items.append((d["n_n"]))
            list_of_items.append(d["rki"])
            list_of_items.append(d["date"])
            list_of_items.append(d["name_lp"])
            list_of_items.append(d["organization"])
            list_of_items.append(d["country"])
            list_of_items.append(d["organization_lp"])
            list_of_items.append(d["date_start"])
            list_of_items.append(d["date_end"])
            list_of_items.append(d["n_protokol"])
            list_of_items.append(d["protokol"])
            list_of_items.append(d["phase"])
            list_of_items.append(d["view"])
            list_of_items.append(d["n_orgs"])
            list_of_items.append(d["n_patient"])
            list_of_items.append(d["type"])
            list_of_items.append(d["status"])
            # list_of_items.append(func(d["clinc_listsl"]))
            list_of_items.append((d["clinc_listsl"][i]))
            list_of_items.append(d["cells"])
            if d["rki"] != "":
                sheet.append(list_of_items)

            list_of_items = []

    r = len(sheet['A'])
    print("Create Sheet_row")

    if new == False:
        sheet_d = book["Sheet_row"]
        sheet = book["Sheet"]
        sheet.insert_rows(idx=2, amount=r)

        for ind, row in enumerate(sheet_d.rows):
            for col, k in enumerate(row):
                sheet.cell(row=ind + 2, column=col + 1).value = k.value

    book.create_sheet("KI")
    sheet_ki = book["KI"]

    print("Create KI")

    for inx, row in enumerate(sheet.rows):
        for cl, t in enumerate(row):
            sheet_ki.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_ki.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_ki.row_dimensions[inx + 1].height = 60

        # if step == 2:
        #     sheet.row_dimensions[row_enumerate].height = 30
        #     sheet.row_dimensions[row_enumerate + 1].height = 30

    print("Alignment 1 iter KI")
    bar = Bar('Processing', max=len(sheet_ki['A']))

    checker = sheet_ki.cell(row=2, column=19).value
    checker2 = sheet_ki.cell(row=2, column=2).value
    for liner, i in enumerate(range(2, len(sheet_ki['A']) + 1)):
        if sheet_ki.cell(row=i, column=19).value == checker and sheet_ki.cell(row=i, column=2).value == checker2:
            pass
        else:

            for col in range(1, 18):
                sheet_ki.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                     end_column=col)
            checker = sheet_ki.cell(row=i, column=19).value
            checker2 = sheet_ki.cell(row=2, column=2).value

        bar.next()

    bar.finish()

    for col in range(1, 18):
        sheet_ki.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                             end_column=col)

    print("Alignment2 iter KI")

    bar = Bar('Processing', max=len(sheet_ki['A']))

    for row in sheet_ki.rows:
        for k in row:
            # print(row[0:-1])
            k.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
            k.font = Font(size="9")
        bar.next()
    bar.finish()

    print("Alignment 3 iter KI")
    bar = Bar('Processing', max=len(sheet_ki['A']))
    for row in sheet_ki.rows:
        for k in row[-2:]:
            k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)

        bar.next()
    bar.finish()

    sheet_ki.column_dimensions[get_column_letter(4)].width = 25
    sheet_ki.column_dimensions[get_column_letter(5)].width = 35
    sheet_ki.column_dimensions[get_column_letter(7)].width = 50
    sheet_ki.column_dimensions[get_column_letter(11)].width = 60
    sheet_ki.column_dimensions[get_column_letter(16)].width = 11
    sheet_ki.column_dimensions[get_column_letter(17)].width = 11
    sheet_ki.column_dimensions[get_column_letter(18)].width = 1000

    book.save("Med_list.xlsx")
    print("file was done")


def get_list_of_clinics(url):
    list_of_clinics = []

    if url != "":

        clinics = get_html(url)

        soup = BeautifulSoup(clinics, "lxml")
        try:
            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                for j in temp:
                    if len(j.text) > 5:
                        list_of_clinics.append(j.text.strip())
        except Exception as e:
            list_of_clinics = []

        return list_of_clinics

    else:

        return list_of_clinics


def main():
    flag = True

    full_list_of_dic = []
    k = 1
    lenght_tables = 25

    url_orig = "https://grls.rosminzdrav.ru/CiPermitionReg.aspx?PermYear=0&DateBeg=&DateEnd=&DateInc=&NumInc=&RegNm=&Statement=&Protocol=&Qualifier=&ProtoNum=&idCIStatementCh=&CiPhase=&RangeOfApp=&Torg=&LFDos=&Producer=&Recearcher=&sponsorCountry=&MedBaseCount=&CiType=&PatientCount=&OrgDocOut=2&Status=1&NotInReg=0&All=0&PageSize={}&order=date_perm&orderType=desc&pagenum=".format(
        lenght_tables)

    result_check = excel_check()

    if result_check == False:

        try:
            while True:
                url = url_orig + str(k)
                html = get_html(url)
                list_of_dics = get_info_for_each_page(html)
                full_list_of_dic.extend(list_of_dics)
                k = k + 1
        except Exception as e:

            print(e, k)

        excel_creator(full_list_of_dic, True)


    else:
        print("Scanning part of data...", str(result_check[0]), str(result_check[1]))

        while flag:

            url = url_orig + str(k)
            # print(url)
            html = get_html(url)
            list_of_dics = get_info_for_each_page(html)
            full_list_of_dic.extend(list_of_dics)

            for item, dic in enumerate(full_list_of_dic):
                if dic["rki"] == result_check[0] and dic["date"] == result_check[1]:
                    append_list_of_dics = full_list_of_dic[0:item + 1]
                    # print(append_list_of_dics[-1])
                    flag = False
                    break

            k = k + 1

        if len(append_list_of_dics) == 1:
            print("No date for udate")
        else:
            excel_creator(append_list_of_dics[:-1], False)


if __name__ == "__main__":
    start = datetime.datetime.now()
    main()
    end = datetime.datetime.now()
    print("Duration: {}".format(end - start))
    input("\nPRESS any batton to complete.....")
