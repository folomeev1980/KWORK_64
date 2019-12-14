import requests
from bs4 import BeautifulSoup
import datetime
from selenium import webdriver
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from progress.bar import Bar
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def create_csv():
    try:

        with open("doctors.csv", "w", newline='') as f:
            pass
    except Exception as e:
        pass
        # print(e)


def writer_csv(data):
    try:
        for dic in data:
            for j in dic["doctors_list"]:
                with open("doctors.csv", "a", newline='') as f:
                    writer = csv.writer(f)
                    a = "{} {} {}".format(dic["famaly"], dic["name"], dic["fathername"])
                    b = "{}_{}_{}".format(j[0], j[1], j[2])
                    writer.writerow((b, a))
    except Exception as e:
        pass
        # print(e)


def read_csv():
    dic = {}
    with open('doctors.csv') as csvfile:
        readCSV = csv.reader(csvfile, delimiter=',')
        for row in readCSV:
            dic[row[0]] = row[1]

    return dic


def get_html(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'}
    r = requests.get(url, headers=headers)
    return r.text


def get_list_of_doctors(url):
    list_of_doctors = []

    if url != "":

        clinics = get_html(url)

        soup = BeautifulSoup(clinics, "lxml")
        try:
            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                a = (temp[-2].text.strip().split("-"))

                b = (temp[-1].text.strip().split()[1][1:])
                list_of_doctors.append([a[0], a[1], b])


        except Exception as e:
            # if e!="'NoneType' object has no attribute 'find_all'":
            #     print(e)
            pass
            list_of_doctors = []

        return list_of_doctors

    else:

        return list_of_doctors


def get_append_list_of_doctors(url):
    list_of_doctors = []
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_argument('window-size=1920x935')
        driver = webdriver.Chrome('chromedriver.exe', options=options)
        driver.get(url)
    except Exception as e:
        # print(e)
        driver.quit()

    k = 2

    while True:

        try:
            # print("ok")
            page = "/html/body/form/table/tbody/tr/td/table/tbody/tr[3]/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr[22]/td/table/tbody/tr/td[{}]/a".format(
                k)
            driver.find_element_by_xpath(page).click()
            html = driver.page_source
            # print(html)

            soup = BeautifulSoup(html, "lxml")

            tds = soup.find("table", class_="ts1").find_all("tr", class_="hi_sys")
            for td in tds:
                temp = td.find_all("td")
                a = (temp[-2].text.strip().split("-"))

                b = (temp[-1].text.strip().split()[1][1:])
                list_of_doctors.append([a[0], a[1], b])

            k = k + 1

        except Exception as e:
            # print("End pages",e)
            break
            # driver.quit()
    driver.quit()

    return list_of_doctors


def get_info_for_each_page(html):
    list_dics = []
    soup = BeautifulSoup(html, "lxml")

    list_of_page = (soup.find_all("tr", class_="hi_sys poi"))

    for st, page in enumerate(list_of_page):

        dic = {}
        list_some = []

        try:
            doctor_link = "https://grls.rosminzdrav.ru/" + \
                          page.get("onclick").split("'")[1]

            doctors_list = get_list_of_doctors(doctor_link)



        except:
            doctors_list = []

        try:

            page_data = page.find_all("td")

            for td in page_data:
                list_some.append(td.text.strip())

            # print(list_some[6])

            if int(list_some[6]) <= 20:
                list_some.append(doctors_list)
            else:
                list_some.append(doctors_list)
                # print(doctor_link)
                list_some[-1].extend(get_append_list_of_doctors(doctor_link))
                # print(list_some)
            # print(list_some)

            dic["famaly"] = list_some[2]
            dic["name"] = list_some[3]
            dic["fathername"] = list_some[4]
            dic["numer_of_ki"] = list_some[6]
            dic["doctors_list"] = list_some[8]

            list_dics.append(dic)


        except Exception as e:
            pass
            # print(e)
            dic = {}
            list_dics.append(dic)

    return list_dics


def excel_edit(dic):
    book = load_workbook("Med_list.xlsx", read_only=True)
    book_new = Workbook()
    sheet = book["Sheet"]
    book_new.create_sheet("KI_Edit")
    sheet_new = book_new["KI_Edit"]

    for inx, row in enumerate(sheet.rows):
        for cl, t in enumerate(row):
            sheet_new.cell(row=inx + 1, column=cl + 1).value = t.value
        if sheet_new.cell(row=inx + 1, column=cl + 1).value == 1:
            sheet_new.row_dimensions[inx + 1].height = 60

    sheet_new.insert_cols(19)
    sheet_new.insert_cols(20)

    reg = ["Онколог", "онколог", "Патолог", "патолог"]

    for inx, row in enumerate(sheet_new.rows):

        s = "{}_{}_{}".format(sheet_new.cell(row=inx + 1, column=8).value,
                              sheet_new.cell(row=inx + 1, column=9).value,
                              sheet_new.cell(row=inx + 1, column=2).value)
        # print(s)
        try:
            sheet_new.cell(row=inx + 1, column=19).value = dic[s]
        except Exception:
            sheet_new.cell(row=inx + 1, column=19).value = ""
        for j in reg:
            if j in str(sheet_new.cell(row=inx + 1, column=11).value) + " " + str \
                        (sheet_new.cell(row=inx + 1, column=16).value):
                sheet_new.cell(row=inx + 1, column=20).value = pacient(
                    str(sheet_new.cell(row=inx + 1, column=11).value))

    #####

    print("Alignment 1 iter KI")
    bar = Bar('Processing', max=len(sheet_new['A']))

    checker = int(sheet_new.cell(row=2, column=21).value)
    checker2 = int(sheet_new.cell(row=2, column=2).value)
    print((checker,checker2))

    for liner, i in enumerate(range(2, len(sheet_new['A']) + 1)):

        if int(sheet_new.cell(row=i, column=21).value) == checker and int(sheet_new.cell(row=i, column=2).value) == checker2:
            # print(i, checker)
            #print("pass")
            pass
        else:
            print(liner)
            for col in range(1, 18):
                # print(i, checker, col)
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            for col in range(19, 21):
                sheet_new.merge_cells(start_row=i - checker, start_column=col, end_row=i - 1,
                                      end_column=col)

            checker = int(sheet_new.cell(row=i, column=21).value)
            checker2 = int(sheet_new.cell(row=i, column=2).value)

        bar.next()

    bar.finish()

    for col in range(1, 18):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)
    for col in range(19, 21):
        sheet_new.merge_cells(start_row=i - checker + 1, start_column=col, end_row=i,
                              end_column=col)

    print("Alignment2 iter KI")

    bar = Bar('Processing', max=len(sheet_new['A']))

    for row in sheet_new.rows:
        for index, k in enumerate(row):
            if index != 17:
                k.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
                k.font = Font(size="9")
            else:
                k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
                k.font = Font(size="9")
        #
        bar.next()
    bar.finish()

    #
    # print("Alignment 3 iter KI")
    # bar = Bar('Processing', max=len(sheet_new['A']))
    # for row in sheet_new.rows:
    #     for k in row[-2:]:
    #         k.alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
    #
    #     bar.next()
    # bar.finish()

    sheet_new.column_dimensions[get_column_letter(4)].width = 25
    sheet_new.column_dimensions[get_column_letter(5)].width = 35
    sheet_new.column_dimensions[get_column_letter(7)].width = 50
    sheet_new.column_dimensions[get_column_letter(11)].width = 60
    sheet_new.column_dimensions[get_column_letter(16)].width = 11
    sheet_new.column_dimensions[get_column_letter(17)].width = 11
    sheet_new.column_dimensions[get_column_letter(18)].width = 200
    sheet_new.column_dimensions[get_column_letter(19)].width = 30
    sheet_new.column_dimensions[get_column_letter(20)].width = 60

    book_new.save("Med_edit_list.xlsx")


def pacient(s):
    l = ["пациент", "Пациент", "Детей", "детей"]

    for i in l:
        ind = s.find(i)
        if ind != -1:
            return "Для " + s[ind:]
        else:
            return ""


def main():
    # i = 1
    # k = 0
    # sum = 0
    # create_csv()
    # while True:
    #     try:
    #         url = "https://grls.rosminzdrav.ru/ciexperts.aspx?F=&N=&P=&D=&ExpertID=&order=fio&orderType=desc&moduleId=2&pageSize=30&pageNum={}".format(
    #             i)
    #         html = get_html(url)
    #         list_of_dics_one_page = get_info_for_each_page(html)
    #         if len(list_of_dics_one_page) == 0:
    #             k = k + 1
    #         sum = sum + len(list_of_dics_one_page)
    #         writer_csv(list_of_dics_one_page)
    #         i = i + 1
    #         print(sum)
    #         if k > 5:
    #             break
    #     except Exception as e:
    #         if k > 5:
    #             break
    # print(i, "Pages\n", sum, "Doctors")

    dic = read_csv()
    print(len(dic))
    excel_edit(dic)


if __name__ == "__main__":
    start = datetime.datetime.now()
    main()
    end = datetime.datetime.now()
    print("Duration: {}".format(end - start))
    input("\nPRESS any batton to complete.....")
