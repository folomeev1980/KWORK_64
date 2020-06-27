from bs4 import BeautifulSoup
from random import choice
import csv
import time
from PIL import Image
from io import BytesIO
import requests
import openpyxl
from openpyxl import Workbook
from rutor import get_proxy
from rutor import get_html


def get_img(url, size=(200, 200)):
    r = requests.get(url, stream=True)
    if not r.ok:
        r.raise_for_error()
    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)


def insert_row(ws, img_url, num0, num1, num2, num3, num4, size=(200, 200)):
    img = openpyxl.drawing.image.Image(get_img(img_url, size=size))
    row_num = ws.max_row + 1
    cell_addr = f"A{row_num}"
    img.anchor = cell_addr
    ws.add_image(img)
    ws[f"B{row_num}"] = num0
    ws[f"C{row_num}"] = num1
    ws[f"D{row_num}"] = num2
    ws[f"E{row_num}"] = num3
    ws[f"F{row_num}"] = num4
    ws[f"G{row_num}"] = "http://rutor.info/search/0/0/000/4/" + num0
    # ws[f"F{row_num}"].value = "link"
    # ws[f"F{row_num}"] = num4
    # row_num[6]=num4
    ws.row_dimensions[row_num].height = int(size[1] * .76)
    ws.column_dimensions["A"].width = int(size[0] * .14)


def write_csv(data):
    with open("kino.csv", "a", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow([data["film_name"],
                         data["year"],
                         "=" + str(data["imdb"]).replace(".", ","),
                         "=" + str(data["kk"]).replace(".", ","),
                         data["cont"],
                         '=ГИПЕРССЫЛКА("{}";"{}")'.format(data["lnk"], "videolnk")])


def repair(lst):
    res = [None, None, None]
    a = []
    for l in lst:
        a.append(l.strip().split(" "))

        res[0] = int(a[0][0])
        if len(a) == 2 and a[1][0] == 'IMDb':
            res[1] = float(a[1][1])
        elif len(a) == 2 and a[1][0] == 'KK':
            res[2] = float(a[1][1])
        elif len(a) > 2 and a[1][0] == 'IMDb' and a[2][0] == 'KK':
            res[1] = float(a[1][1])
            res[2] = float(a[2][1])

    return res


def get_page_data_1(html, ws):
    soup = BeautifulSoup(html, "lxml")
    movies = soup.find("div", id="movies").find("div", id="catalog-content").find_all("div", class_="movie")
    for movie in movies:
        flag = True
        film_name = movie.find("img").get("alt")
        info = movie.find("span", class_="add-info").text.strip().split(",")
        year = repair(info)[0]
        imdb = repair(info)[1]
        kk = repair(info)[2]
        lnk = "https:" + movie.find("img", class_="cover").get("src").strip()
        cont = movie.find_all("span")[-1].text.strip()

        data = {"film_name": film_name,
                "year": year,
                "imdb": imdb,
                "kk": kk,
                "cont": cont,
                "lnk": lnk}
        for i in cont.split(","):
            if i in ["Комедия", "Музыка", "Спорт", "Документальный", "Фэнтези", "Мелодрама"]:
                flag = False
        if flag == True:
            size = (80, 120)
            # wb = Workbook()

            # hiper = '=HYPERLINK("{}";"{}")'.format(
            #     "http://rutor.info/search/0/0/000/4/" + data["film_name"],
            #     "videolnk")

            insert_row(ws, data["lnk"], data["film_name"], data["year"], data["imdb"], data["kk"], data["cont"],
                       size=size)


def main():
    wb = Workbook()
    ws = wb.active
    for i in range(1, 5, 1):
        flag = True
        url = "https://www.kinokopilka.pro/movies?page={}&period=today&sort=released_on".format(str(i))
        while flag == True:
            html = (get_html(url))

            try:
                get_page_data_1(html, ws)

                flag = False
            except:
                flag = True

        print("Page " + str(i) + " parsed")
    wb.save('kino1_.xlsx')


if __name__ == "__main__":
    main()
